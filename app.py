import base64
import io
import os
import re
import time
from typing import List, Optional, Tuple
from urllib.parse import urlparse

import requests
from fastapi import FastAPI, Header, HTTPException
from openpyxl import Workbook
from pydantic import BaseModel

APP_API_KEY = os.getenv("APP_API_KEY", "")
SERPAPI_KEY = os.getenv("SERPAPI_KEY", "")
DEBUG_LOGS = os.getenv("DEBUG_LOGS", "0") == "1"

# blocca directory / terze parti + news/magazine (falsi positivi frequenti)
BLOCKED_DOMAINS = {
    "linkedin.com", "facebook.com", "instagram.com", "x.com", "twitter.com",
    "paginegialle.it", "paginebianche.it", "kompass.com", "europages.com",
    "crunchbase.com", "dnb.com", "bloomberg.com", "reuters.com", "wikipedia.org",
    "google.com", "maps.google.com", "google.it",

    # news / magazine / portali
    "alvolante.it", "quattroruote.it", "ilsole24ore.com", "repubblica.it",
    "corriere.it", "ansa.it", "it.wikipedia.org", "wikipedia.it",
}

# token "pericolosi" (troppo generici) -> NON usare fallback primary.it/com
COMMON_TOKENS = {
    "star", "next", "punto", "nuova", "geo", "tech", "lli", "innovation",
    "system", "group", "holding", "service", "services", "italy", "italia",
    "engineering", "meccanica", "elettronica", "mediterranea", "cold",
    "dama", "vba", "mps", "re", "home", "spa", "srl",
}

# Throttling SerpAPI (stabilità)
SERPAPI_QPS_DELAY = 1.0   # pausa tra chiamate SerpAPI (secondi)
SERPAPI_429_SLEEP = 25    # pausa lunga se 429
SERPAPI_TIMEOUT = 25

app = FastAPI()


class EnrichRequest(BaseModel):
    companies: List[str]
    output_filename: Optional[str] = "domini_compilati.xlsx"


def _log(msg: str) -> None:
    if DEBUG_LOGS:
        print(msg, flush=True)


def normalize_domain(d: str) -> str:
    d = d.strip().lower().replace("www.", "")
    if "://" in d:
        d = urlparse(d).netloc.lower().replace("www.", "")
    d = d.split("/")[0]
    d = re.sub(r"[^a-z0-9\.\-]", "", d)
    return d


def to_root_domain(dom: str) -> str:
    dom = normalize_domain(dom)
    parts = dom.split(".")
    if len(parts) <= 2:
        return dom
    two_level_tlds = {"co.uk", "com.au", "co.jp", "com.br", "com.tr"}
    last2 = ".".join(parts[-2:])
    last3 = ".".join(parts[-3:])
    if last2 in two_level_tlds and len(parts) >= 3:
        return last3
    return last2


def is_blocked(domain: str) -> bool:
    dom = normalize_domain(domain)
    for b in BLOCKED_DOMAINS:
        if dom == b or dom.endswith("." + b):
            return True
    return False


def tld_score_bonus(root_domain: str) -> float:
    d = normalize_domain(root_domain)
    if d.endswith(".com"):
        return 0.18
    if d.endswith(".it"):
        return 0.15
    if d.endswith(".org"):
        return 0.05
    if d.endswith(".net"):
        return -0.10
    if d.endswith(".info"):
        return -0.25
    return 0.0


def serpapi_search(company: str, mode: str = "it") -> List[str]:
    """
    Ricerca SerpAPI con retry/backoff + throttling.
    mode:
      - "it": base
      - "it_deep": più profondo (solo se serve)
      - "en": fallback inglese (solo se serve)
    """
    if not SERPAPI_KEY:
        raise RuntimeError("SERPAPI_KEY non configurata")

    if mode == "en":
        queries = [f"\"{company}\" official website"]
        hl, gl = "en", "us"
    elif mode == "it_deep":
        queries = [
            f"\"{company}\" \"P.IVA\"",
            f"\"{company}\" \"partita IVA\"",
            f"\"{company}\" \"contatti\"",
        ]
        hl, gl = "it", "it"
    else:
        queries = [
            f"\"{company}\" sito ufficiale",
            f"\"{company}\" contatti",
        ]
        hl, gl = "it", "it"

    urls: List[str] = []
    for q in queries:
        params = {
            "engine": "google",
            "q": q,
            "hl": hl,
            "gl": gl,
            "num": 10,
            "api_key": SERPAPI_KEY,
        }

        # backoff più lungo, soprattutto su 429
        backoffs = [0, 2, 8, 20]
        last_err = None

        for wait_s in backoffs:
            if wait_s:
                time.sleep(wait_s)

            try:
                _log(f"[SERPAPI] q={q}")
                r = requests.get("https://serpapi.com/search.json", params=params, timeout=SERPAPI_TIMEOUT)

                # throttling tra chiamate SerpAPI
                time.sleep(SERPAPI_QPS_DELAY)

                if r.status_code == 429:
                    _log("[SERPAPI] 429 rate limit -> sleep")
                    last_err = RuntimeError("SerpAPI HTTP 429")
                    time.sleep(SERPAPI_429_SLEEP)
                    continue

                if r.status_code in (500, 502, 503, 504):
                    _log(f"[SERPAPI] {r.status_code} temporary -> retry")
                    last_err = RuntimeError(f"SerpAPI HTTP {r.status_code}")
                    time.sleep(3)
                    continue

                r.raise_for_status()
                data = r.json()
                for item in (data.get("organic_results") or []):
                    link = item.get("link")
                    if link:
                        urls.append(link)

                last_err = None
                break

            except Exception as e:
                _log(f"[SERPAPI] exception: {e}")
                last_err = e
                continue

        if last_err is not None:
            raise last_err

    seen = set()
    urls = [u for u in urls if not (u in seen or seen.add(u))]
    return urls


def fetch_text_multi(root_domain: str, timeout_s: int, deep: bool) -> Optional[str]:
    """
    base: homepage
    deep: homepage + pagine corporate tipiche
    """
    def get(url: str) -> Optional[str]:
        try:
            resp = requests.get(url, timeout=timeout_s, headers={"User-Agent": "Mozilla/5.0"})
            return (resp.text or "").lower()
        except Exception:
            return None

    base_urls = [f"https://{root_domain}", f"http://{root_domain}"]
    paths = [""] + (["/contatti", "/chi-siamo", "/about", "/privacy", "/contacts", "/company"] if deep else [])

    collected: List[str] = []
    for base in base_urls:
        ok_any = False
        for p in paths:
            t = get(base + p)
            if t:
                collected.append(t)
                ok_any = True
        if ok_any:
            break

    if not collected:
        return None
    return ("\n".join(collected))[:250000]


def tokenize_company(company: str) -> Tuple[List[str], str]:
    raw_tokens = [t for t in re.split(r"\W+", company.lower()) if len(t) >= 3]
    stop = {"spa", "s", "p", "a", "s.p.a", "s.p.a.", "srl", "s.r.l", "s.r.l.", "societa", "società", "company", "group", "holding"}
    tokens = [t for t in raw_tokens if t not in stop][:3]
    primary = tokens[0] if tokens else ""
    return tokens, primary


def is_safe_fallback_token(primary: str) -> bool:
    if not primary or len(primary) < 4:
        return False
    if primary in COMMON_TOKENS:
        return False
    return True


def score_domain(company: str, tokens: List[str], primary: str, dom_root: str, text: str) -> float:
    score = 0.0
    score += tld_score_bonus(dom_root)

    if primary and primary in dom_root:
        score += 0.6

    token_hits = sum(1 for tok in tokens if tok and tok in text)
    if token_hits >= 2:
        score += 0.6
    elif token_hits == 1:
        score += 0.25

    if ("partita iva" in text) or ("p.iva" in text) or ("codice fiscale" in text) or ("vat" in text):
        score += 0.35
    if ("contatti" in text) or ("chi siamo" in text) or ("about" in text):
        score += 0.1
    if ("cookie" in text) or ("privacy" in text):
        score += 0.05

    if ("news" in text) or ("newsletter" in text) or ("articolo" in text) or ("abbonati" in text):
        score -= 0.35
    if ("directory" in text) or ("scheda azienda" in text):
        score -= 0.5

    if len(tokens) <= 1 and primary and (primary not in dom_root):
        if not (("partita iva" in text) or ("p.iva" in text) or ("vat" in text) or (company.lower() in text)):
            score -= 0.6

    return score


def pick_best_domain(company: str) -> Tuple[str, float]:
    start = time.time()
    MAX_SECONDS_PER_COMPANY = 35  # Profilo A stabile

    def time_left() -> bool:
        return (time.time() - start) <= MAX_SECONDS_PER_COMPANY

    tokens, primary = tokenize_company(company)

    # fallback utile per ST
    if "stmicroelectronics" in company.lower() or "st microelectronics" in company.lower():
        t = fetch_text_multi("st.com", timeout_s=8, deep=False)
        if t:
            return "st.com", 0.95

    def evaluate(urls: List[str], deep_fetch: bool, max_candidates: int) -> Tuple[str, float]:
        candidates: List[str] = []
        for u in urls:
            host = normalize_domain(u)
            if not host:
                continue
            root = to_root_domain(host)
            if not root or is_blocked(root):
                continue
            candidates.append(root)

        seen = set()
        candidates = [c for c in candidates if not (c in seen or seen.add(c))]

        best = ("NON TROVATO", 0.0)
        for dom_root in candidates[:max_candidates]:
            if not time_left():
                break
            text = fetch_text_multi(dom_root, timeout_s=12, deep=deep_fetch)
            if not text:
                continue
            sc = score_domain(company, tokens, primary, dom_root, text)
            if sc > best[1]:
                best = (dom_root, sc)
            if best[1] >= 0.92:
                break
        return best

    # PASS 1: IT base
    try:
        urls_it = serpapi_search(company, mode="it")
    except Exception as e:
        _log(f"[PICK] SerpAPI IT failed: {e}")
        urls_it = []
    best_dom, best_score = evaluate(urls_it, deep_fetch=False, max_candidates=4)

    # fallback token sicuro primary.it/com (deep fetch)
    if best_score < 0.80 and is_safe_fallback_token(primary) and time_left():
        for tld in ("it", "com"):
            guess = to_root_domain(f"{primary}.{tld}")
            if is_blocked(guess):
                continue
            text = fetch_text_multi(guess, timeout_s=10, deep=True)
            if not text:
                continue
            sc = score_domain(company, tokens, primary, guess, text)
            if sc >= 0.82:
                return guess, sc

    # PASS 2: IT deep solo se serve
    if (best_dom == "NON TROVATO" or best_score < 0.82) and time_left():
        try:
            urls_deep = serpapi_search(company, mode="it_deep")
        except Exception as e:
            _log(f"[PICK] SerpAPI IT_DEEP failed: {e}")
            urls_deep = []
        dom2, sc2 = evaluate(urls_deep, deep_fetch=True, max_candidates=6)
        if sc2 > best_score:
            best_dom, best_score = dom2, sc2

    # PASS 3: EN fallback solo se ancora basso
    if (best_dom == "NON TROVATO" or best_score < 0.82) and time_left():
        try:
            urls_en = serpapi_search(company, mode="en")
        except Exception as e:
            _log(f"[PICK] SerpAPI EN failed: {e}")
            urls_en = []
        dom3, sc3 = evaluate(urls_en, deep_fetch=True, max_candidates=5)
        if sc3 > best_score:
            best_dom, best_score = dom3, sc3

    if best_score < 0.82:
        return "NON TROVATO", best_score

    return to_root_domain(best_dom), best_score


def require_bearer_token(authorization: str | None) -> None:
    if not APP_API_KEY:
        raise HTTPException(status_code=500, detail="APP_API_KEY non configurata sul server")
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Unauthorized")
    token = authorization.split(" ", 1)[1].strip()
    if token != APP_API_KEY:
        raise HTTPException(status_code=401, detail="Unauthorized")


@app.post("/enrich/domains")
def enrich_domains(req: EnrichRequest, authorization: str | None = Header(default=None)):
    require_bearer_token(authorization)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Ragione sociale"
    ws["B1"] = "Dominio"

    for i, name in enumerate(req.companies, start=2):
        ws[f"A{i}"] = name  # NON modificare

        try:
            domain, _score = pick_best_domain(name)
        except Exception as e:
            _log(f"[ROW] error on '{name}': {e}")
            domain = "NON TROVATO"

        ws[f"B{i}"] = domain if domain else "NON TROVATO"

        # stabilità su run lunghi (anti-rate-limit e anti-burst)
        time.sleep(0.6)

    buf = io.BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode("utf-8")

    return {
        "openaiFileResponse": [
            {
                "name": req.output_filename or "domini_compilati.xlsx",
                "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "content": b64,
            }
        ]
    }


@app.get("/health")
def health():
    return {"ok": True}
